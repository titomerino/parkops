
from io import BytesIO
from django.utils import timezone
from openpyxl import Workbook
from django.http import HttpResponse
from django.template.loader import render_to_string
import weasyprint

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)


def minutes_to_hours_and_minutes(total_minutes: int):
    hours = total_minutes // 60
    minutes = total_minutes % 60

    return f"{hours:02d}", f"{minutes:02d}"


def format_plate(plate: str) -> str:
    """
    Formatea una placa así:
    - Primera letra sola
    - El resto agrupado de derecha a izquierda en bloques de 3

    Ej:
    P40807  -> P 40 807
    P8E98   -> P 8 E98
    P911116 -> P 911 116
    """

    if not plate:
        return ""

    plate = plate.strip().upper()

    if len(plate) <= 1:
        return plate

    first = plate[0]
    rest = plate[1:]

    # Calcular tamaño del primer grupo (lo que sobra al dividir entre 3)
    remainder = len(rest) % 3

    groups = []

    if remainder:
        groups.append(rest[:remainder])

    for i in range(remainder, len(rest), 3):
        groups.append(rest[i:i+3])

    return f"{first} {' '.join(groups)}"


def render_pdf_response(
    request,
    template_name,
    context,
    filename,
):
    html_string = render_to_string(
        template_name,
        context
    )

    pdf = weasyprint.HTML(
        string=html_string,
        base_url=request.build_absolute_uri("/")
    ).write_pdf()

    response = HttpResponse(
        pdf,
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        f'inline; filename="{filename}"'
    )

    return response


def export_report_excel(context, report_date, report_end_date=None, type=None, plate=None):

    wb = Workbook()

    ws = wb.active
    ws.title = "Reporte diario"

    # ESTILOS
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="212529"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    center = Alignment(
        horizontal="center"
    )

    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # TOTAL
    ws["E1"] = "Total ingresos"
    ws["E1"].font = bold_font

    total_cell = ws["F1"]

    total_cell.value = float(context["total_income"])
    total_cell.font = bold_font
    total_cell.number_format = "$#,##0.00"


    # DETALLE
    row = 3

    headers = [
        "Placa",
        "Entrada",
        "Salida",
        "Tiempo",
        "Tipo",
        "Monto",
    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=row,
            column=col,
            value=header
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    row += 1

    # DATOS
    for entry in context["entries"]:

        ws.cell(
            row=row,
            column=1,
            value=entry.plate
        )

        entry_date = timezone.localtime(entry.entry_date_hour)

        ws.cell(
            row=row,
            column=2,
            value=entry_date.strftime(
                "%d/%m - %I:%M %p"
            )
            if entry_date else ""
        )

        departure_date = timezone.localtime(entry.departure_date_hour)

        ws.cell(
            row=row,
            column=3,
            value=departure_date.strftime(
                "%d/%m - %I:%M %p"
            )
            if departure_date else ""
        )

        ws.cell(
            row=row,
            column=4,
            value=entry.duration
        )

        ws.cell(
            row=row,
            column=5,
            value=entry.type + " - " + entry.fee.name if entry.fee else entry.type
        )

        amount_cell = ws.cell(
            row=row,
            column=6,
            value=float(entry.final_amount or 0)
        )

        amount_cell.number_format = "$#,##0.00"

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border

        row += 1

    # AUTOAJUSTE COLUMNAS
    for column in ws.columns:

        max_length = 0

        try:
            column_letter = column[0].column_letter
        except AttributeError:
            continue

        for cell in column:

            if cell.value:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            column_letter
        ].width = max_length + 3


    # DESCARGA
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    report_title = "reporte-parqueo-"

    match type:
        case "day":
            report_title = report_title + "diario-" + report_date.strftime("%d-%m-%Y") + ".xlsx"
        case "monthly":
            report_title = report_title + "mensual-" + report_date.strftime("%m-%Y") + ".xlsx"
        case "period":
            report_title = report_title + "periodo-" + report_date.strftime("%d-%m-%Y") + "-al-" + report_end_date.strftime("%d-%m-%Y") + ".xlsx"
        case "plate":
            report_title = report_title + "placa-" + plate + "-" + report_date.strftime("%d-%m-%Y") + "-al-" + report_end_date.strftime("%d-%m-%Y") + ".xlsx"
        case _:
                return HttpResponse("Tipo de reporte no válido", status=400)

    response["Content-Disposition"] = (
        f'attachment; filename={report_title}'
    )

    return response